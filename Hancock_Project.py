# Hancock Project (Only Condos and Heat Exchanger 2)
# Xhoi Shyti *** Anchor Mechanical
#
#
# This program calculates averages based on measurements from various HVAC systems in the Hancock Building
# Takes a CSV file as input and writes the result for each room to a new CSV file


# openpyxl is the third party module that allows the program to read and intepret excel files.
# times is the custom module that allows the program to treat the timestamps as integer values so that they can be compared to the billing schedule.
# convert_Excel is the custom module that takes the incoming CSV and converts to an Excel format.
# outputCSV is the custom module that converts back to CSV.

import openpyxl
import datetime
import times
import convert_Excel
from outputCSV import toCSV

# Condos are billed 24/7.
# Thus, BTUs are calculated for each row in the file, regardless of the timestamp.
def condos():

    # Convert CSV to excel, then load to a workbook object
    # load the current, active sheet
    condos = convert_Excel.convert_CONDOS('Report_2.csv')
    condos_sheet = condos.active

    #Initialize the empty arrays 
    BTUs = []
    averages = [] 
    
    # Iterate through the all the rows and apply the designated formula that calculates BTU
    # Add the BTUs to the array
    for row in range(2, condos_sheet.max_row + 1):
        BTUs.append(((float(condos_sheet.cell(row = row, column = 3).value) - float(condos_sheet.cell(row = row, column = 2).value)) * 499 * float(condos_sheet.cell(row = row, column = 4).value)) / 1000)
    
    # Modify the BTUs array so that it is comprised of "chunks" (subarrays) that are 4 elements wide
    # 4 elements because the timestamps are in 15 miniute intervals, and Anchor Mechanical is concerned with the hourly average.
    chunks = [BTUs[x:x+4] for x in range(0, len(BTUs), 4)]
    
    # Calculate the average of each hour and add the result to the averages array
    for group in chunks:
        averages.append(sum(group)/len(group))
    
    totalSum = sum(averages)
    
    # Print the result on the command line, and write the result to a new CSV file
    print("Condos:")
    print(totalSum)
    row = ["Condos", totalSum]
    toCSV(row)


def he2():
    he2 = convert_Excel.convert_HE2('Report.csv')
    he2_sheet = he2.active
    
    BTUs = []
    averages = []
    
    # Unlike condos, Heat Exchanger 2 has a custom billing schedule for every day of the week.
    # Thus, BTUs are calcuated only if the timestamp fits in the schedule (Correct weekday and correct hour).
    for row in range(2, he2_sheet.max_row + 1):
        if times.convertDateTime(he2_sheet.cell(row = row, column = 1)).date().weekday() == 6:
            if times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] >= 0 and times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] < 1000:
                BTUs.append(((float(he2_sheet.cell(row = row, column = 3).value) - float(he2_sheet.cell(row = row, column = 2).value)) * 499 * float(he2_sheet.cell(row = row, column = 4).value)) / 1000)
		
        elif times.convertDateTime(he2_sheet.cell(row = row, column = 1)).date().weekday() == 0:
            if times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] >= 0 and times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] < 800:
                BTUs.append(((float(he2_sheet.cell(row = row, column = 3).value) - float(he2_sheet.cell(row = row, column = 2).value)) * 499 * float(he2_sheet.cell(row = row, column = 4).value)) / 1000)
				
        elif times.convertDateTime(he2_sheet.cell(row = row, column = 1)).date().weekday() == 1:
            if times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] >= 0 and times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] < 800:
                BTUs.append(((float(he2_sheet.cell(row = row, column = 3).value) - float(he2_sheet.cell(row = row, column = 2).value)) * 499 * float(he2_sheet.cell(row = row, column = 4).value)) / 1000)
				
        elif times.convertDateTime(he2_sheet.cell(row = row, column = 1)).date().weekday() == 2:
            if times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] >= 0 and times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] < 800:
                BTUs.append(((float(he2_sheet.cell(row = row, column = 3).value) - float(he2_sheet.cell(row = row, column = 2).value)) * 499 * float(he2_sheet.cell(row = row, column = 4).value)) / 1000)
				
        elif times.convertDateTime(he2_sheet.cell(row = row, column = 1)).date().weekday() == 3:
            if times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] >= 0 and times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] < 800:
                BTUs.append(((float(he2_sheet.cell(row = row, column = 3).value) - float(he2_sheet.cell(row = row, column = 2).value)) * 499 * float(he2_sheet.cell(row = row, column = 4).value)) / 1000)
				
        elif times.convertDateTime(he2_sheet.cell(row = row, column = 1)).date().weekday() == 4:
            if times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] >= 100 and times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] < 800:
                BTUs.append(((float(he2_sheet.cell(row = row, column = 3).value) - float(he2_sheet.cell(row = row, column = 2).value)) * 499 * float(he2_sheet.cell(row = row, column = 4).value)) / 1000)
				
        elif times.convertDateTime(he2_sheet.cell(row = row, column = 1)).date().weekday() == 5:
            if times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] >= 100 and times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] < 800:
                BTUs.append(((float(he2_sheet.cell(row = row, column = 3).value) - float(he2_sheet.cell(row = row, column = 2).value)) * 499 * float(he2_sheet.cell(row = row, column = 4).value)) / 1000)
            elif times.easyTime[str(he2_sheet.cell(row = row, column = 1).value)[-15:]] >= 2200:
                BTUs.append(((float(he2_sheet.cell(row = row, column = 3).value) - float(he2_sheet.cell(row = row, column = 2).value)) * 499 * float(he2_sheet.cell(row = row, column = 4).value)) / 1000)
    
    chunks = [BTUs[x:x+4] for x in range(0, len(BTUs), 4)]
    
    for group in chunks:
        averages.append(sum(group)/len(group))

    totalSum = sum(averages)

    print("HE2:")
    print(totalSum)
    row = ["HE2", totalSum]
    toCSV(row)
